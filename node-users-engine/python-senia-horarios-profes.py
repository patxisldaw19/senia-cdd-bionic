#!/usr/bin/env python3
#
#

# Dependencias : python3-openpyxl


import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook
import openpyxl


print(" * Welcome to python horarios")

class Sesion:

    def __init__(self,dia_semana,sesion_orden,aula,grupo,materia):
        self.dia_semana = dia_semana
        self.sesion_orden = sesion_orden
        self.aula = aula
        self.grupo = grupo
        self.materia = materia

    def print_sesion(self):
        print(" * "+self.dia_semana+" : "+self.sesion_orden+" - "+self.grupo+" - "+self.aula)


class Profesor:

    def __init__(self,nombre,apellido):
        self.nombre=nombre
        self.apellido = apellido
        self.listaSesiones = []

# Cargamos el fichero de origen
horarios_file="patatadecolores.xlsx"
horarios_profes="patatasfritas.xlsx"

wb_orig = load_workbook(filename = horarios_file)
ws_orig = wb_orig.active

wb_dest = Workbook()

profesores = []


## Metodos para trabajar con los profesores

def busca_y_anyade(nombre,apellido):
    aux_profe = Profesor(nombre,apellido)
    found = False
    for profe in profesores:
        if profe.nombre == aux_profe.nombre and profe.apellido == aux_profe.apellido:
            found = True
            return profe
    if not found:
        profesores.append(aux_profe)
        return aux_profe


## 
for columna in range(2,100):
    for fila in range(2,100):
        celda_aux = ws_orig.cell(column=columna,row=fila)
        if (celda_aux.value == None):
            pass
        else:
            print("--------------")
            #print(celda_aux.value)

            # Trabajamos esta celda
            hora = ws_orig.cell(column=1,row=fila).value.strip('\r\n')

            # Dirty hack
            dia_hora = hora[:1]
            hora_hora = hora.split(":",1)[1]
            
            datos_celda = celda_aux.value.replace("\n"," ").split(" ")
            print(datos_celda)

            ## Dirty hack?
            p_nombre_aux = datos_celda[2]
            if p_nombre_aux == "Mª":
                p_nombre_aux += datos_celda[3]
                p_apellidos_aux = datos_celda[4]
            else:
                p_apellidos_aux = datos_celda[3]

            p_grupo = datos_celda[0]
            p_materia = datos_celda[1]
            
            aula = ws_orig.cell(column=columna,row=1).value.rstrip()

            print(dia_hora+":"+hora_hora+" -> "+aula)

            profe_aux = busca_y_anyade(p_nombre_aux,p_apellidos_aux)

            sesion_aux = Sesion(dia_hora,hora_hora,aula,p_grupo,p_materia)
            profe_aux.listaSesiones.append(sesion_aux)
            

# Ahora rellenamos el excel
for profe in profesores:
    print("++++++++++++++++++")
    print(" * "+profe.nombre+" . "+ profe.apellido+":"+str(len(profe.listaSesiones)))

    # Create sheet
    ws_dest_profe = wb_dest.create_sheet(profe.nombre+"."+profe.apellido)

    # Profesor
    celda=ws_dest_profe.cell(column=1,row=1)
    celda.value=profe.nombre+"."+profe.apellido
    celda.style='Headline 1'

    ws_dest_profe.column_dimensions["A"].width=25

    for col in ["B","C","D","E","F"]:
        ws_dest_profe.column_dimensions[col].width=15
        ws_dest_profe.column_dimensions[col].alignment = Alignment(horizontal='center')
        ws_dest_profe.column_dimensions[col].alignment = Alignment(vertical='center')

    for row in range(1,20):
        ws_dest_profe.row_dimensions[row].height=40
        
    
    # Horas
    columna=1
    fila=2
    for hora in range(1,19):
        celda = ws_dest_profe.cell(column=columna,row=fila)
        celda.value=fila-1
        celda.style='Headline 2'
        fila=fila+1

    # Dias
    fila=2
    columna=2
    for dia in ["L","M","X","J","V"]:
        celda = ws_dest_profe.cell(column=columna,row=1)
        celda.value=dia

        celda.style='Headline 1'
        celda.alignment=Alignment(vertical="center")
        celda.alignment=Alignment(horizontal="center")

    
        for sesion in profe.listaSesiones:
            print(sesion.dia_semana+","+sesion.sesion_orden+","+sesion.aula+","+sesion.grupo+","+sesion.materia)
            if (sesion.dia_semana==dia):
                celda_sesion=ws_dest_profe.cell(column=columna,row=int(sesion.sesion_orden)+1)
                celda_sesion.value="["+sesion.aula+"]\n"+sesion.grupo+":\n"+sesion.materia

        # Incrementamos la columna
        columna=columna+1

wb_dest.save("patatasfritas.xlsx")




